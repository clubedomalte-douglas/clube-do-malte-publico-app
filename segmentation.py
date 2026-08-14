"""
Motor de segmentacao de publico - Clube do Malte / Caixa da Alegria

Implementa o processo padronizado definido com o Douglas em 13/08/2026:

FASE 1 (produto com historico de vendas):
  1) Ja comprou o produto exato antes
  2) E ja comprou produtos de estilo similar antes (fora do proprio produto)
  3) E fez 3+ pedidos avulsos no total (historico geral)
  4) E NAO comprou o produto exato nos ultimos 30 dias
  -> cruza com telefone valido + opt-in WhatsApp = 'Sim'
  -> ordena por Score RFV equilibrado (percentil Recencia+Frequencia+Valor, peso 1/3 cada)

FASE 2 (fallback, acionado quando o publico elegivel da Fase 1 fica abaixo de um limite,
        tipico em lancamentos sem historico):
  Camada A - Estilo especifico similar (prioridade): mesma logica do criterio 2 da Fase 1,
             sem exigir ter comprado o produto exato. Ainda exige 3+ pedidos totais e nao
             ter comprado nada da familia nos ultimos 30 dias. Ranqueada por RFV.
  Camada B - Familia/categoria generica (so preenche o que a Camada A nao cobrir): usa a
             coluna de familia oficial do arquivo de pedidos (LEVES, LUPULADAS, FRUTADAS,
             MALTADAS, COMPLEXAS, TORRADAS, AZEDAS, Sem Alcool, KITs).
"""

import re
import unicodedata
import numpy as np
import pandas as pd

FAMILIAS = ["LEVES", "LUPULADAS", "FRUTADAS", "MALTADAS", "COMPLEXAS",
            "TORRADAS", "AZEDAS", "Sem Álcool", "KITs"]

SPOT_COLS = ['CD_CADASTRO', 'ID_Pedido', 'Cidade', 'UF', 'CEP', 'Data Pedido',
             'Data Primeira Compra', 'Data Ultima Compra', 'Valor do Pedido',
             'Valor do Frete', 'Valor dos Produtos', 'LEVES', 'LUPULADAS', 'FRUTADAS',
             'MALTADAS', 'COMPLEXAS', 'TORRADAS', 'AZEDAS', 'Sem Álcool', 'KITs',
             'Produtos Internos', 'SKU Interno', 'Telefone', 'Email', 'Nome', 'CPF']

LEADSCORE_COLS = ['Email', 'Nome', 'Telefone', 'CPF', 'Cidade', 'UF', 'Lead_Score',
                   'Pontos_Visita_Pagina', 'Pontos_Compra', 'Pontos_Assinatura_Ativa',
                   'Ja_Foi_Assinante', 'Assinante_Ativo_Hoje', 'Qtd_Pedidos_Avulsos',
                   'Receita_Total_Avulsos', 'Data_Ultima_Atividade',
                   'Dias_Desde_Ultima_Atividade', 'Opt_Email_Netdeal', 'Opt_SMS_Netdeal',
                   'Opt_Whatsapp_Netdeal', 'Fontes', 'Segmento', 'Microbase_WAPP']

LEADSCORE_SHEETS = ['01_ASS_Alto_Score_VIP', '02_ASS_Score_Medio', '03_EXASS_Winback_Quente',
    '04_EXASS_Winback_Morno_90dd_1an', '05_EXASS_Winback_Morno_Acima1an', '06_EXASS_Winback_Frio',
    '07_SPOT_Comprador_Quente', '08_SPOT_Comprador_Morno', '09_SPOT_NaoComprador_EngQuente',
    '10_SPOT_NaoComprador_EngMorno', '11_SPOT_NaoComprador_EngFrio']


def _to_num(s):
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return np.nan
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip().replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return np.nan


def _valid_phone(p):
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return False
    digits = ''.join(ch for ch in str(p) if ch.isdigit())
    return len(digits) >= 10


def _fmt_phone_whatsapp(p):
    digits = ''.join(ch for ch in str(p) if ch.isdigit())
    if len(digits) in (10, 11):
        digits = '55' + digits
    return digits


def _norm_keywords(raw):
    if not raw:
        return []
    parts = [p.strip() for p in re.split(r'[,;]', raw) if p.strip()]
    out = []
    for p in parts:
        p = unicodedata.normalize('NFKD', p).encode('ascii', 'ignore').decode('ascii').lower()
        out.append(p)
    return out


def _norm_text_series(s):
    s = s.astype(str).str.lower()
    s = s.apply(lambda x: unicodedata.normalize('NFKD', x).encode('ascii', 'ignore').decode('ascii'))
    return s


def load_spot_pedidos(file_like):
    """Leitura rapida via openpyxl read_only (pd.read_excel e muito lento em arquivos >20MB)."""
    import openpyxl
    wb = openpyxl.load_workbook(file_like, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    n = len(SPOT_COLS)
    rows = ws.iter_rows(min_row=2, max_col=n, values_only=True)
    header = next(rows)
    # se a 2a linha for uma linha de descricao (nao dados), pula
    data = []
    first = next(rows, None)
    if first is not None:
        joined = ' '.join(str(x) for x in first if x is not None).lower()
        if not any(k in joined for k in ['produto', 'pedido', 'cliente']) or (first[0] not in (None, '') and str(first[0]).replace('.', '').isdigit()):
            data.append(first)
    for row in rows:
        if row[0] is None:
            continue
        row = list(row) + [None] * (n - len(row))
        data.append(row[:n])
    wb.close()
    df = pd.DataFrame(data, columns=SPOT_COLS)

    for c in ['Valor do Pedido', 'Valor do Frete', 'Valor dos Produtos'] + FAMILIAS:
        df[c + '_num'] = df[c].apply(_to_num)

    df['Data Pedido_dt'] = pd.to_datetime(df['Data Pedido'], errors='coerce')
    df['Email_norm'] = df['Email'].astype(str).str.strip().str.lower()
    df['Produtos_norm'] = _norm_text_series(df['Produtos Internos'].fillna(''))
    return df


def load_leadscore(file_like):
    """Leitura rapida via openpyxl read_only."""
    import openpyxl
    wb = openpyxl.load_workbook(file_like, read_only=True, data_only=True)
    n = len(LEADSCORE_COLS)
    frames = []
    for sname in LEADSCORE_SHEETS:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        data = []
        for row in ws.iter_rows(min_row=2, max_col=n, values_only=True):
            if row[0] is None:
                break
            row = list(row) + [None] * (n - len(row))
            data.append(row[:n])
        frames.append(pd.DataFrame(data, columns=LEADSCORE_COLS))
    wb.close()
    master = pd.concat(frames, ignore_index=True)
    master['Email_norm'] = master['Email'].astype(str).str.strip().str.lower()
    master = master.drop_duplicates(subset='Email_norm', keep='first')
    return master


def load_rd_station(file_like):
    """
    Le o export de leads da RD Station (Email, Nome, Telefone, Datas de conversao,
    Eventos etc). Usa 'Data da ultima conversao' como proxy de ultima navegacao/
    engajamento no site, pois o export nao traz timestamp por evento de navegacao.
    Aceita .csv (utf-16 ou utf-8, separador ; ou ,) ou .xlsx.
    """
    name = getattr(file_like, 'name', str(file_like))
    if str(name).lower().endswith(('.xlsx', '.xls')):
        df = pd.read_excel(file_like)
    else:
        raw_bytes = file_like.read() if hasattr(file_like, 'read') else open(file_like, 'rb').read()
        df = None
        for enc in ('utf-16', 'utf-8', 'latin1'):
            try:
                import io as _io
                df = pd.read_csv(_io.BytesIO(raw_bytes), sep=None, engine='python', encoding=enc)
                if df.shape[1] > 1:
                    break
            except Exception:
                df = None
        if df is None or df.shape[1] <= 1:
            raise ValueError('Nao foi possivel ler o arquivo da RD Station (encoding/separador nao reconhecido).')

    col_email = next((c for c in df.columns if c.strip().lower() == 'email'), None)
    col_ultima = next((c for c in df.columns if 'ltima convers' in c.lower()), None)
    if col_email is None or col_ultima is None:
        raise ValueError("Arquivo da RD Station precisa ter as colunas 'Email' e 'Data da ultima conversao'.")

    out = df[[col_email, col_ultima]].rename(columns={col_email: 'Email', col_ultima: 'Data_Ultima_Navegacao_RD'})
    out['Email_norm'] = out['Email'].astype(str).str.strip().str.lower()
    out['Data_Ultima_Navegacao_RD'] = pd.to_datetime(
        out['Data_Ultima_Navegacao_RD'].astype(str).str.replace(r'\s*[-+]\d{4}$', '', regex=True),
        errors='coerce')
    out = out.dropna(subset=['Data_Ultima_Navegacao_RD'])
    out = out.sort_values('Data_Ultima_Navegacao_RD').drop_duplicates(subset='Email_norm', keep='last')
    return out[['Email_norm', 'Data_Ultima_Navegacao_RD']]


def aggregate_customers(spot_df, exact_keywords, similar_keywords, familia, ref_date=None):
    if ref_date is None:
        ref_date = pd.Timestamp.today().normalize()

    df = spot_df.copy()
    txt = df['Produtos_norm']

    exact_pat = '|'.join(re.escape(k) for k in exact_keywords) if exact_keywords else None
    similar_pat = '|'.join(re.escape(k) for k in similar_keywords) if similar_keywords else None

    df['flag_exact'] = txt.str.contains(exact_pat, regex=True, na=False) if exact_pat else False

    if similar_pat:
        txt_no_exact = txt
        for k in exact_keywords:
            txt_no_exact = txt_no_exact.str.replace(re.escape(k), '', regex=True)
        df['flag_similar_row'] = txt_no_exact.str.contains(similar_pat, regex=True, na=False)
    else:
        df['flag_similar_row'] = False

    fam_col = familia + '_num' if familia in FAMILIAS else None

    base = df.groupby('Email_norm').agg(
        Qtd_Pedidos=('ID_Pedido', 'nunique'),
        Valor_Total=('Valor dos Produtos_num', 'sum'),
        Data_Ultima_Compra=('Data Pedido_dt', 'max'),
        Qtd_Pedidos_Exato=('flag_exact', 'sum'),
        flag_similar_any=('flag_similar_row', 'any'),
        Nome=('Nome', 'first'),
        Telefone=('Telefone', 'first'),
        Cidade=('Cidade', 'first'),
        UF=('UF', 'first'),
    )

    exact_rows = df[df['flag_exact']]
    last_exact = exact_rows.groupby('Email_norm')['Data Pedido_dt'].max().rename('Data_Ultima_Compra_Exato')

    if fam_col:
        non_exact = df[~df['flag_exact']]
        fam_qty = non_exact.groupby('Email_norm')[fam_col].sum().rename('Familia_qty_non_exact')
    else:
        fam_qty = pd.Series(dtype=float, name='Familia_qty_non_exact')

    agg = base.join(last_exact, how='left').join(fam_qty, how='left').reset_index()
    agg['Familia_qty_non_exact'] = agg['Familia_qty_non_exact'].fillna(0)

    agg['Ja_Comprou_Exato'] = agg['Qtd_Pedidos_Exato'] > 0
    agg['Ja_Comprou_Similar_Estilo'] = agg['flag_similar_any']
    agg['Ja_Comprou_Familia'] = agg['Familia_qty_non_exact'] > 0

    agg['Dias_Desde_Ultima_Compra'] = (ref_date - agg['Data_Ultima_Compra']).dt.days
    agg['Dias_Desde_Ultima_Compra_Exato'] = (ref_date - agg['Data_Ultima_Compra_Exato']).dt.days

    return agg


def _add_rfv(pool):
    pool = pool.copy()
    pool['R_score'] = 100 - pool['Dias_Desde_Ultima_Compra'].rank(pct=True) * 100
    pool['F_score'] = pool['Qtd_Pedidos'].rank(pct=True) * 100
    pool['V_score'] = pool['Valor_Total'].rank(pct=True) * 100
    pool['RFV_score'] = (pool['R_score'] + pool['F_score'] + pool['V_score']) / 3
    return pool


def _cross_eligibility(pool, lead_master):
    lead_sub = lead_master[['Email_norm', 'Telefone', 'Opt_Whatsapp_Netdeal', 'Segmento',
                             'Lead_Score', 'Assinante_Ativo_Hoje']].rename(
        columns={'Telefone': 'Telefone_LS'})
    m = pool.merge(lead_sub, on='Email_norm', how='left')
    m['Telefone_final'] = m['Telefone_LS'].fillna(m['Telefone'])
    m['has_phone'] = m['Telefone_final'].apply(_valid_phone)
    eligible = m[m['has_phone'] & (m['Opt_Whatsapp_Netdeal'] == 'Sim')].copy()
    return eligible


def gerar_publico(spot_df, lead_master, exact_keywords_raw, similar_keywords_raw, familia,
                   tamanho=400, min_pedidos=3, dias_exclusao=30, limite_fallback=100,
                   ref_date=None, rd_df=None, dias_navegacao=360):
    exact_kw = _norm_keywords(exact_keywords_raw)
    similar_kw = _norm_keywords(similar_keywords_raw)

    agg = aggregate_customers(spot_df, exact_kw, similar_kw, familia, ref_date=ref_date)
    funil = [("Base total de clientes com pedidos", len(agg))]

    c1 = agg['Ja_Comprou_Exato']
    funil.append(("Criterio 1 - comprou o produto exato", int(c1.sum())))
    c12 = c1 & agg['Ja_Comprou_Similar_Estilo']
    funil.append(("Criterio 1 E 2 - + estilo similar", int(c12.sum())))
    c123 = c12 & (agg['Qtd_Pedidos'] >= min_pedidos)
    funil.append((f"Criterio 1 E 2 E 3 - + {min_pedidos}+ pedidos totais", int(c123.sum())))
    c1234 = c123 & ((agg['Dias_Desde_Ultima_Compra_Exato'] > dias_exclusao) | agg['Dias_Desde_Ultima_Compra_Exato'].isna())
    funil.append((f"Criterio 1 E 2 E 3 E 4 - + nao comprou nos ult. {dias_exclusao}d", int(c1234.sum())))

    fase1_pool = agg[c1234].copy()
    fase1_elig = _cross_eligibility(fase1_pool, lead_master)
    funil.append(("Fase 1 - elegiveis (telefone + opt-in WhatsApp)", len(fase1_elig)))
    fase1_elig = _add_rfv(fase1_elig)
    fase1_elig['Origem'] = 'Fase 1 - Produto exato'

    resultado = fase1_elig.sort_values('RFV_score', ascending=False)
    modo = 'fase1'

    if len(fase1_elig) < limite_fallback:
        modo = 'fallback_camada_a'
        camada_a_pool = agg[
            agg['Ja_Comprou_Similar_Estilo'] &
            (agg['Qtd_Pedidos'] >= min_pedidos) &
            ((agg['Dias_Desde_Ultima_Compra_Exato'] > dias_exclusao) | agg['Dias_Desde_Ultima_Compra_Exato'].isna())
        ].copy()
        camada_a_elig = _cross_eligibility(camada_a_pool, lead_master)
        funil.append(("Fallback Camada A - estilo similar, elegiveis", len(camada_a_elig)))
        camada_a_elig = _add_rfv(camada_a_elig)
        camada_a_elig['Origem'] = 'Fallback A - Estilo similar'

        combinado = pd.concat([fase1_elig, camada_a_elig], ignore_index=True)
        combinado = combinado.drop_duplicates(subset='Email_norm', keep='first')
        combinado = combinado.sort_values(['Origem', 'RFV_score'], ascending=[True, False])
        resultado = combinado

        if len(combinado) < tamanho and familia in FAMILIAS:
            modo = 'fallback_camada_a_b'
            camada_b_pool = agg[
                agg['Ja_Comprou_Familia'] &
                (agg['Qtd_Pedidos'] >= min_pedidos) &
                ((agg['Dias_Desde_Ultima_Compra_Exato'] > dias_exclusao) | agg['Dias_Desde_Ultima_Compra_Exato'].isna())
            ].copy()
            camada_b_elig = _cross_eligibility(camada_b_pool, lead_master)
            funil.append((f"Fallback Camada B - familia '{familia}', elegiveis", len(camada_b_elig)))
            camada_b_elig = _add_rfv(camada_b_elig)
            camada_b_elig['Origem'] = 'Fallback B - Familia'

            combinado2 = pd.concat([combinado, camada_b_elig], ignore_index=True)
            combinado2 = combinado2.drop_duplicates(subset='Email_norm', keep='first')

            ordem_origem = {'Fase 1 - Produto exato': 0, 'Fallback A - Estilo similar': 1, 'Fallback B - Familia': 2}
            combinado2['ordem'] = combinado2['Origem'].map(ordem_origem)
            combinado2 = combinado2.sort_values(['ordem', 'RFV_score'], ascending=[True, False]).drop(columns='ordem')
            resultado = combinado2

    # ---------------- Criterio extra: navegacao recente (RD Station) ----------------
    if rd_df is not None and len(rd_df):
        resultado = resultado.merge(rd_df, on='Email_norm', how='left')
        resultado['Dias_Desde_Ultima_Navegacao_RD'] = (
            (ref_date if ref_date is not None else pd.Timestamp.today().normalize())
            - resultado['Data_Ultima_Navegacao_RD']
        ).dt.days
        resultado['Navegou_Recente'] = resultado['Dias_Desde_Ultima_Navegacao_RD'] <= dias_navegacao
        funil.append((f"+ Navegaram no site nos ultimos {dias_navegacao}d (RD Station)",
                      int(resultado['Navegou_Recente'].sum())))
    else:
        resultado['Navegou_Recente'] = np.nan
        resultado['Dias_Desde_Ultima_Navegacao_RD'] = np.nan

    if 'Origem' in resultado.columns:
        ordem_origem = {'Fase 1 - Produto exato': 0, 'Fallback A - Estilo similar': 1, 'Fallback B - Familia': 2}
        resultado['_ordem_origem'] = resultado['Origem'].map(ordem_origem).fillna(0)
    else:
        resultado['_ordem_origem'] = 0

    if rd_df is not None and len(rd_df):
        resultado['_ordem_nav'] = (~resultado['Navegou_Recente'].fillna(False)).astype(int)
        resultado = resultado.sort_values(
            ['_ordem_origem', '_ordem_nav', 'RFV_score'], ascending=[True, True, False])
    else:
        resultado = resultado.sort_values(['_ordem_origem', 'RFV_score'], ascending=[True, False])
    resultado = resultado.drop(columns=[c for c in ['_ordem_origem', '_ordem_nav'] if c in resultado.columns])

    final = resultado.head(tamanho).copy()
    final['Telefone (WhatsApp)'] = final['Telefone_final'].apply(_fmt_phone_whatsapp)
    final['RFV_score'] = final['RFV_score'].round(1)
    final['Valor_Total'] = final['Valor_Total'].round(2)
    final = final.reset_index(drop=True)
    final.insert(0, '#', range(1, len(final) + 1))

    cols_out = ['#', 'Nome', 'Telefone (WhatsApp)', 'Email_norm', 'Cidade', 'UF', 'Segmento',
                'Origem', 'Qtd_Pedidos_Exato', 'Qtd_Pedidos', 'Valor_Total',
                'Dias_Desde_Ultima_Compra', 'Navegou_Recente', 'Dias_Desde_Ultima_Navegacao_RD',
                'Assinante_Ativo_Hoje', 'RFV_score']
    for c in cols_out:
        if c not in final.columns:
            final[c] = np.nan
    final = final[cols_out].rename(columns={
        'Email_norm': 'Email',
        'Qtd_Pedidos_Exato': 'Qtd. pedidos c/ produto exato',
        'Qtd_Pedidos': 'Qtd. pedidos totais',
        'Valor_Total': 'Valor total comprado (R$)',
        'Dias_Desde_Ultima_Compra': 'Dias desde ultima compra',
        'Navegou_Recente': f'Navegou no site (ult. {dias_navegacao}d)',
        'Dias_Desde_Ultima_Navegacao_RD': 'Dias desde ultima navegacao (RD)',
        'Assinante_Ativo_Hoje': 'Assinante ativo hoje',
        'RFV_score': 'Score RFV (0-100)',
    })

    return {'tabela': final, 'funil': funil, 'modo': modo}
